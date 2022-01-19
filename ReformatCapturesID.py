import numpy as np
from numpy.lib.shape_base import _column_stack_dispatcher
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import load_workbook
import openpyxl as pxl
import eel
from pandas.io import excel
import shutil

eel.init('web', allowed_extensions=['.js', '.html'])

# directory = input("Enter The Directory of All of the Files (EX: C:\\edang\\Desktop): ")
# os.chdir("G:\\Shared drives\\Process Development\\TAM\\AE_Method 3 to 7\\Emily Tam_ReTest Auto_Iggy Wheatley\\Iggy_Wheatley\\qPCR_AE6_Trblsh_Wheatley\\ExtravaganzaFile(pls_ignore)")
# os.chdir(directory)

# filename = input("Enter File Name (.csv at the end): ")

# filename = input("123-456_202107261703.csv")
# excelfilename = input("Enter Excel File Name (.xlsx at the end): ")
# excelfilename = input("AndysTemplate_Auto_qPCR_Wheatley_v1.2.xlsx")
@eel.expose
def Captureto96WellFormat(file,qPCRfile):
    outputFile = pd.read_csv(file)
    CaptureID = outputFile['Capture_ID'].unique()
    numSamples = len(CaptureID)

    column1 = ["Neg Control"]
    column2 = ["Pos Control"]
    column3 = []
    column5 = []
    column7 = []
    column9 = []
    column11 = []

    dataFormat = pd.DataFrame(np.zeros((8, 12)), index = ["A","B","C","D","E","F","G","H"])
    dataFormat.columns = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']

    for x,y in zip(range(numSamples),CaptureID):
        if x <= 6:
            column1.append(y)
            column2.append(y)
        if x >6 and x<=14:
            column3.append(y)
        if x>14 and x<=22:
            column5.append(y)
        if x>22 and x<=30:
            column7.append(y)
        if x>30 and x<=38:
            column9.append(y)
        if x>38 and x<=46:
            column11.append(y)

    columns = [column1, column2, column3, column3, column5, column5, column7, column7, column9, column9,column11, column11]

    for i, column in enumerate(columns):
        if len(column) != 8:
            # print(8 - len(column))
            
            dataFormat[str(i+ 1)] = (column) + (np.zeros((1, 8 - len(column))) * np.nan).ravel().tolist()
            # print(dataFormat[str(i+ 1)])                                        
        
        else:
            dataFormat[str(i+ 1)] = column
            # print(dataFormat[str(i+ 1)])
    print(dataFormat)

    df_qPCRdata = pd.read_csv(qPCRfile,skiprows=19)
    df_holder = pd.read_csv(qPCRfile,nrows=18)
    df2_qPCRdata = df_qPCRdata.sort_values(by=['Content'],ascending=True)
    frames=[df_holder,df2_qPCRdata]
    result = pd.concat(frames,ignore_index=True,sort=False)

    excelfile = "G:\\Shared drives\\Process Development\\TAM\\AE_Method 3 to 7\\Auto_qPCR_Template_ED_10212021_Final.xlsx"
    excel_book = pxl.load_workbook(excelfile)
    with pd.ExcelWriter(excelfile, engine="openpyxl") as writer:
        writer.book = excel_book
        writer.sheets = {
            worksheet.title: worksheet
            for worksheet in excel_book.worksheets
        }
        DF = pd.DataFrame(dataFormat)
        DF.to_excel(writer, "CaptureLayout", index=False)
        result.to_excel(writer, "Copy data here",index=False)
        writer.save()    

    root = tk.Tk()
    root.attributes('-topmost', 1)
    root.withdraw()
    filePath = filedialog.asksaveasfile(mode='a', defaultextension = '.xlsx')
    root.destroy()
    shutil.copy(excelfile,filePath.name)

    print("Complete!!!")
    return dataFormat

@eel.expose
def df2excelpd(excelfilename,dataFormat):
    excel_book = pxl.load_workbook(excelfilename)
    with pd.ExcelWriter(excelfilename, engine="openpyxl") as writer:
        writer.book = excel_book
        writer.sheets = {
            worksheet.title: worksheet
            for worksheet in excel_book.worksheets
        }
        DF = pd.DataFrame(dataFormat)
        DF.to_excel(writer, "CaptureLayout", index=False)
        writer.save()
    print("Complete!!")

@eel.expose
def df2excel(dataFormat):
    dataFormat.to_excel("CaptureID")

@eel.expose
def dfExcel(outputDf,filePath):
    df = pd.DataFrame(outputDf)
    df.to_excel(filePath.name)
    print(df)
    return filePath, "Complete!!!"



